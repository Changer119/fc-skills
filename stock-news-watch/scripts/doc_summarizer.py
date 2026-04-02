# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "httpx>=0.27",
#     "pypdf>=4.0",
#     "python-docx>=1.1",
#     "openpyxl>=3.1",
# ]
# ///
"""下载披露文件并使用 AI 模型生成摘要。"""

from __future__ import annotations

import logging
from dataclasses import dataclass

import httpx

logger = logging.getLogger(__name__)

MAX_TEXT_LENGTH = 8000  # 发给 AI 的最大文本长度


@dataclass
class LLMConfig:
    """大模型配置。"""

    api_url: str
    api_token: str
    model: str


@dataclass
class SummaryResult:
    """摘要结果。"""

    success: bool
    summary: str = ""
    error: str = ""


async def download_file(url: str, timeout: float = 60.0) -> bytes | None:
    """下载文件并返回字节内容。"""
    try:
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        logger.error("下载文件失败 %s: %s", url, e)
        return None


def extract_text_from_pdf(content: bytes) -> str:
    """从 PDF 提取文本。"""
    from pypdf import PdfReader
    import io

    try:
        reader = PdfReader(io.BytesIO(content))
        texts: list[str] = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
        return "\n".join(texts)
    except Exception as e:
        logger.error("解析 PDF 失败: %s", e)
        return ""


def extract_text_from_word(content: bytes) -> str:
    """从 Word 文档提取文本。"""
    from docx import Document
    import io

    try:
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        logger.error("解析 Word 失败: %s", e)
        return ""


def extract_text_from_excel(content: bytes) -> str:
    """从 Excel 提取文本。"""
    from openpyxl import load_workbook
    import io

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        texts: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception as e:
        logger.error("解析 Excel 失败: %s", e)
        return ""


def extract_text_from_html(content: bytes) -> str:
    """从 HTML 提取纯文本。"""
    import re

    try:
        text = content.decode("utf-8", errors="ignore")
        text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text
    except Exception as e:
        logger.error("解析 HTML 失败: %s", e)
        return ""


EXTRACTORS = {
    "pdf": extract_text_from_pdf,
    "word": extract_text_from_word,
    "excel": extract_text_from_excel,
    "html": extract_text_from_html,
}


def extract_text(content: bytes, file_type: str) -> str:
    """根据文件类型提取文本。"""
    extractor = EXTRACTORS.get(file_type, extract_text_from_pdf)
    return extractor(content)


async def call_llm(config: LLMConfig, prompt: str) -> str:
    """调用兼容 OpenAI 接口的大模型。"""
    headers = {
        "Authorization": f"Bearer {config.api_token}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": config.model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "你是一个专业的金融分析助手。请用简洁的中文总结以下股票信息披露文件的关键内容。"
                    "重点关注：财务数据变化、重大事项、风险提示、业绩预告等。"
                    "输出格式：先给出一句话概要，然后列出 3-5 个要点。"
                ),
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.3,
        "max_tokens": 2000,
    }

    url = config.api_url.rstrip("/")
    if not url.endswith("/chat/completions"):
        url = f"{url}/chat/completions"

    async with httpx.AsyncClient(timeout=600.0) as client:
        resp = await client.post(url, headers=headers, json=payload)
        if resp.status_code != 200:
            logger.error(
                "LLM API 返回 %d: %s", resp.status_code, resp.text[:500]
            )
            resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"]


async def summarize_disclosure(
    file_url: str,
    file_type: str,
    title: str,
    stock_name: str,
    llm_config: LLMConfig,
) -> SummaryResult:
    """下载文件并生成 AI 摘要。

    Args:
        file_url: 文件下载地址
        file_type: 文件类型 (pdf/word/excel/html)
        title: 公告标题
        stock_name: 股票名称
        llm_config: 大模型配置
    """
    # 下载文件
    content = await download_file(file_url)
    if content is None:
        return SummaryResult(success=False, error="文件下载失败")

    # 提取文本
    text = extract_text(content, file_type)
    if not text.strip():
        return SummaryResult(
            success=True,
            summary=f"【{stock_name}】{title}\n（文件内容无法提取文本，可能为扫描件或图片）",
        )

    # 截断过长文本
    if len(text) > MAX_TEXT_LENGTH:
        text = text[:MAX_TEXT_LENGTH] + "\n...(内容过长，已截断)"

    # 调用 AI 摘要
    try:
        prompt = f"股票：{stock_name}\n公告标题：{title}\n\n公告正文：\n{text}"
        summary = await call_llm(llm_config, prompt)
        return SummaryResult(success=True, summary=summary)
    except Exception as e:
        import traceback
        err_type = type(e).__name__
        logger.error("AI 摘要失败 [%s]: %s\n%s", err_type, e, traceback.format_exc())
        return SummaryResult(success=False, error=f"AI 摘要失败 [{err_type}]: {e}")
